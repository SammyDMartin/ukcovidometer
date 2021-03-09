import os
import requests
import json
import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as st
from html.parser import HTMLParser
import ast
import re
from tqdm import tqdm
from scipy.optimize import minimize as minmagic
from IPython.display import display
import pandas as pd
import datetime
import csv
import warnings
import formulas
from pdfreader import SimplePDFViewer
import openpyxl

import os, sys

class HiddenPrints:
    def __enter__(self):
        self._original_stdout = sys.stdout
        sys.stdout = open(os.devnull, 'w')

    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stdout.close()
        sys.stdout = self._original_stdout

warnings.filterwarnings("ignore")

STEP = 7
uk_pop = 68.1*10**6
england_pop = 55.98*10**6
USEPERCENT = False

def moving_average(a, n) :
    ret = np.cumsum(a, dtype=float)
    ret[n:] = ret[n:] - ret[:-n]
    return ret[n - 1:] / n

def c_moving_average(x, w):
    return np.convolve(x, np.ones(w), 'valid') / w

def extract_clean_finaldata(filename):
    projections=pd.read_excel(filename)
    #data = df.parse(sheet_name='deaths')
    #print(data.type)
    ensemble = {}
    models = ['NoLockdown','OpenSchools','OpenSchoolsWorst','200k','2M','LockdownWorst']
    for model in models:
        sample = projections[[model+"_dates",model+"_hos"]]

        model_dates = sample[model+'_dates'].dropna()
        model_deaths = sample[model+'_hos'].dropna().tolist()

        sample = pd.Series(np.array(model_deaths), index = model_dates)

        sample = sample.resample('D').mean().interpolate(method='polynomial', order=3)

        deaths = sample.tolist()
        dates = sample.index.tolist()

        for idx,d in enumerate(dates):
            dates[idx] = d.date()

        ensemble[model] = [dates,deaths]
        
    return ensemble


def extract_clean_doomdata(filename):
    projections=pd.read_excel(filename)
    #data = df.parse(sheet_name='deaths')
    #print(data.type)
    ensemble = {}
    models = ['insane','high','med1','med2','worstcase']
    for model in models:
        sample = projections[[model+"_dates",model+"_deaths"]]

        model_dates = sample[model+'_dates'].dropna()
        model_deaths = sample[model+'_deaths'].dropna().tolist()

        sample = pd.Series(np.array(model_deaths), index = model_dates)

        sample = sample.resample('D').mean().interpolate(method='polynomial', order=3)

        deaths = sample.tolist()
        dates = sample.index.tolist()

        for idx,d in enumerate(dates):
            dates[idx] = d.date()

        ensemble[model] = [dates,deaths]
        
    return ensemble



def produce_datetimes(datelist):
    return [datetime.datetime.strptime(date,'%Y-%m-%d') for date in datelist]

def produce_timestring(date):
    return date.strftime("%Y-%m-%d")






def spew_text(url,start,end):
    proxies = {}
    response = requests.get(url = url, proxies = proxies)
    print("STATUS:", response.status_code)
    data = str(response.text)

    start = data.find(start)
    end = data[start:].find(end)
    end = start+end

    print(start,end)
    data = data[start:end]
    return data

def get_json(url):
    proxies = {}
    response = requests.get(url = url, proxies = proxies)
    #print("STATUS:", response.status_code)
    return response.json()

def diagprint(dates,data):
    for i in range(len(data)):
        print(i,dates[i].strftime("%m-%d"),int(data[i]))


def positive(name,url, ax,col,avg_per):
    
    hospital = get_json(url)['data']

    dates_h,data_h = [],[]

    for x in range(len(hospital)):
        dates_h.append(hospital[x]['date'])
        data_h.append(hospital[x]['uniqueCasePositivityBySpecimenDateRollingSum'])

    if avg_per > 1:
        dates_h,data_h = dates_h[::-1],data_h[::-1]

        data_h = c_moving_average(data_h,avg_per)

        data_h = [round(d,1) for d in data_h]

        dates_h = dates_h[avg_per-1:]

        dates_h,data_h = dates_h[::-1],data_h[::-1]        

    squished = [date[5:] for date in dates_h[:4]]

    titlestr = "Positivity rate in {} by date:\n{}\n{}".format(name,squished,data_h[:4])
    if name == "South East":
        print(titlestr)

    dates_h = produce_datetimes(dates_h)

    ax.plot(dates_h,data_h,col+'--',linewidth=1.0,label=titlestr)

def hospitalsadmissions(name,url, ax,col):
    
    hospital = get_json(url)['data']

    dates_h,data_h = [],[]

    for x in range(len(hospital)):
        dates_h.append(hospital[x]['date'])
        data_h.append(hospital[x]['newAdmissions'])

    
    max_wave1 = max(data_h[-100:])

    squished = [date[5:] for date in dates_h[:4]]
    change = data_h[:4]

    factor = np.array(change)/max_wave1
    factor = list(factor)
    factor = [round(f,2) for f in factor]
    
    dates_h,data_h = dates_h[::-1],data_h[::-1]
    
    data_h = moving_average(data_h,7)
    dates_h = dates_h[6:]

    #data_h = new(data_h)
    #dates_h = dates_h[1:]

    titlestr = "Admissions {} on date:\n{}\n{}".format(name,squished,change)

    if name == "South East":
        print(titlestr)

    dates_h = produce_datetimes(dates_h)
    ax.plot(dates_h,np.array(data_h),col,linewidth=1.0,label=titlestr)


def hospitalschange(name,url, ax,col):
    
    hospital = get_json(url)['data']

    dates_h,data_h = [],[]

    for x in range(len(hospital)):
        dates_h.append(hospital[x]['date'])
        data_h.append(hospital[x]['hospitalCases'])

    
    max_wave1 = max(data_h[-100:])

    squished = [date[5:] for date in dates_h[:4]]
    change = data_h[:4]

    factor = np.array(change)/max_wave1
    factor = list(factor)
    factor = [round(f,2) for f in factor]
    
    dates_h,data_h = dates_h[::-1],data_h[::-1]
    
    data_h = moving_average(data_h,7)
    dates_h = dates_h[6:]

    data_h = new(data_h)
    dates_h = dates_h[1:]

    #data_h = growth_rate(data_h)
    #dates_h = dates_h[2:]

    #for idx in range(len(data_h))[1:]:
    #    data_h[idx] = data_h[idx] - data_h[idx-1]
    
    #dates_h,data_h = dates_h[1:],data_h[1:]


    """
    out = []
    for idx in range(len(data_h))[1:]:
        out.append(data_h[idx]-data_h[idx-1])

    dates_h = dates_h[1:]
    data_h = out

    """


    titlestr = "In H {} on date:\n{}\n{}\n(current/first wave:)\n{}".format(name,squished,change,factor)

    if name == "South East":
        print(titlestr)

    dates_h = produce_datetimes(dates_h)


    ax.plot(dates_h,np.array(data_h),col,linewidth=1.5,label=titlestr)

def cases(name,url, ax,col,avg_per):
    
    hospital = get_json(url)['data']

    dates_h,data_h = [],[]

    for x in range(len(hospital)):
        dates_h.append(hospital[x]['date'])
        data_h.append(hospital[x]['newCasesBySpecimenDateChangePercentage'])


    if avg_per > 1:
        dates_h,data_h = dates_h[::-1],data_h[::-1]

        data_h = c_moving_average(data_h,avg_per)

        data_h = [round(d,1) for d in data_h]

        dates_h = dates_h[avg_per-1:]

        dates_h,data_h = dates_h[::-1],data_h[::-1]        

    squished = [date[5:] for date in dates_h[:4]]

    titlestr = "Growth rate in {} by date:\n{}\n{}".format(name,squished,data_h[:4])
    if name == "South East":
        print(titlestr)

    dates_h = produce_datetimes(dates_h)

    ax.plot(dates_h,data_h,col,linewidth=1.5,label=titlestr)

def ascended_approximation(win_ratio, slow, fast, van_morrison_value):
    fit = np.polyfit(np.array([2e5,2e6]),np.array([0,1]), 1)

    fit_fn = np.poly1d(fit)

    frac = fit_fn(win_ratio*2e6)


    if van_morrison_value is not None:
        start = datetime.date(2020,12,20)
        end = datetime.date(2021,3,1)
    else:
        start = datetime.date(2021,2,1)
        end = datetime.date(2021,5,1)

    endpoints = lambda tup,start,end: [tup[0].index(start), tup[0].index(end)]

    dates = [slow[0],None]
    
    slow = slow[1][endpoints(slow,start,end)[0]:endpoints(slow,start,end)[1]]
    fast = fast[1][endpoints(fast,start,end)[0]:endpoints(fast,start,end)[1]]

    dates = dates[0][endpoints(dates,start,end)[0]:endpoints(dates,start,end)[1]]

    fast,slow = np.array(fast),np.array(slow)

    ensemble = (frac*fast) + ((1-frac)*slow)

    value_on_day = lambda model,day : model[1][model[0].index(day)]

    if van_morrison_value == None:
        return  [dates,ensemble]

    dates = [d + datetime.timedelta(6) for d in dates]
    prediction = [dates,ensemble]

    initial = value_on_day(prediction,datetime.date(2020,12,26))
    
    upshift = van_morrison_value - initial
    
    ensemble = ensemble+upshift

    prediction = [dates,ensemble]

    return prediction



def SE_hospitals(per, win_ratio):
    #ST HOSPITALS
    """
    
    . We compared three main scenarios for non-pharmaceutical interventions: (i)a counterfactual scenario with Tiers 1–3 only, i.e. without additional Tier 4 restrictions thatwere first introduced on 20 December 2020; (ii)
    Tier 4 introduced from 20 December 2020 inEast of England, London, and the South East, with Tier 4 restrictions introduced from 26December 2020 in all other regions of England,
    lasting until 31 January 2021 and with schoolsand universities opening from 4 January 2021; (iii) scenario ii, but with schools and universitiesremaining closed until 31 January 2021.
    We also examined two vaccination scenarios: (iv)200,000 vaccinations per week and (v) 2,000,000 vaccinations per week. Both vaccinationscenarios occurred against a backdrop of non-pharmaceutical
    interventions as in scenario iii.

    Args:
        per (int, optional): [description]. Defaults to 3.
    """
    shift = 2
    hurl = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=South%2520East&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newAdmissions%22:%22newAdmissions%22,%22cumAdmissions%22:%22cumAdmissions%22%7D&format=json"
    hospital = get_json(hurl)['data']

    dates_h,data_h = [],[]

    for x in range(len(hospital)):
        dates_h.append(hospital[x]['date'])
        data_h.append(hospital[x]['newAdmissions'])

    titlestr = "Admissions in SE:\n{}\n{}\n".format(dates_h[:3][::-1],data_h[:3][::-1])
    print(titlestr)

    data_h,dates_h = data_h[::-1],dates_h[::-1]
    #diagprint(dates_h,data_h)

    data_h = moving_average(data_h,per)
    dates_h = dates_h[per-1:]

    wave_1 = data_h[:40]
    first_peak = max(wave_1)
    
    dates_h = produce_datetimes(dates_h)

    fig, ax = plt.subplots(figsize=(15,15))
    #print(titlestr)
    plt.title(titlestr)
    #diagprint(dates_h,data_h)
    ax.plot(dates_h[::-1],data_h[::-1],'k',linewidth=1.5,label='{} day moving average'.format(per))
    #ax.plot(np.arange(len(dates_h)),np.ones(len(dates_h))*maxh,'--')
    #ax.xaxis.set_major_locator(plt.MaxNLocator(10))
    #ax.set_title(out)
    plt.ylabel("Hospital Admissions")

    proj = extract_clean_finaldata('finalstrike.xlsx')
    
    for k in proj.keys():
        val = proj[k]

        m = [None,None]

        dates = val[0][per-1:]

        m[0] = [d - datetime.timedelta(shift) for d in dates]
        m[1] = moving_average(val[1],per)


        labels = ['NoLockdown','OpenSchools','OpenSchoolsWorst','200k','2M','LockdownWorst']

        #diagprint(m[0],m[1])

        value_on_day = lambda model,day : model[1][model[0].index(day)]
        
        if k is labels[0]:
            ax.plot(m[0],m[1],'y-',label="Van Morrison",linewidth=0.5)
            correction = value_on_day(m,datetime.date(2020,12,26))
        elif k is labels[1]:
            ax.plot(m[0],m[1],'r-',label="Schools Open",linewidth=0.5)
        elif k is labels[2]:
            ax.plot(m[0],m[1],'r--',label="Schools Open_worst",linewidth=0.2)
        elif k is labels[3]:
            ax.plot(m[0],m[1],'b-',label="200k Vaccinated",linewidth=1.0)
            k2 = m
        elif k is labels[4]:
            ax.plot(m[0],m[1],'g-',label="2M Vaccinated",linewidth=1.0)
            m2 = m
            bl = m[0]
        elif k is labels[5]:
            ax.plot(m[0],m[1],'g--',label="200k/2M_worst",linewidth=0.2)
    
    model = ascended_approximation(win_ratio,k2,m2,correction)
    model_lite = ascended_approximation(win_ratio,k2,m2,None)
    ax.plot(model[0],model[1],'m-',label="Deep Research UltraApproximation, high_ratio={}".format(round(win_ratio,3)),linewidth=1.5)
    ax.plot(model_lite[0],model_lite[1],'m--',label="Deep Research Approximation, high_ratio={}".format(round(win_ratio,3)),linewidth=1.0)

    textstr = "In all except 'Van Morrison Lockdown Modelled from\n20/12/20 to 31/1/21\nActual 26/12/20 to ??\nSchools Open Models no Vaccinated and opening 04/01/21"

    ax.text(0.05, 0.95, textstr, transform=ax.transAxes, fontsize=12,
        verticalalignment='top')

    plt.plot(bl,first_peak*np.ones(len(bl)),'p--',label='First Peak',linewidth=0.5)
    ax.legend()

    fig.savefig("6 SEhospitals_full.png")

    plt.xlim(datetime.datetime(2020,11,1),datetime.datetime(2021,5,1))

    #plt.show()
    fig.savefig("6 SEhospitals.png")

def OWID_vacc():
    url = 'https://raw.githubusercontent.com/owid/covid-19-data/master/public/data/vaccinations/vaccinations.csv'
    response = requests.get(url = url, proxies = {})
    #print("STATUS:", response.status_code)
    #print("URL:",response.url)
    data = response.content.decode('utf-8')
    cr = csv.reader(data.splitlines(), delimiter=',')
    data = list(cr)
    try:
        v_dates,v_data = [],[]
        v_dates.append(datetime.datetime(2020,12,8))
        v_data.append(0)
        for row in data:
            if row[0] == 'United Kingdom' and (row[3] is not ""):
                date = datetime.datetime.strptime(row[2],'%Y-%m-%d')
                data = float(row[3])
                v_dates.append(date)
                v_data.append(data)

        print("Owid Vacc:")
        for idx,i in enumerate(v_dates):
            print(produce_timestring(v_dates[idx]),v_data[idx])

        return v_dates,v_data
    except Exception as E:
        for row in data:
            if row[0] == 'United Kingdom':
                print(row)
        print("Max Roser being funny")
        return [],[]

def Graph_region(dataframe,loclist,axlist,name):
    locs = ','.join(loclist)
    data = dataframe.parse(sheet_name='1f',skiprows=7,usecols = "A,"+locs)

    ax = axlist[0]
    ax_diff = axlist[1]
    
    stream = data.values.tolist()
    dates,means,lows,highs = [],[],[],[]
    for content in stream:
        if str(type(content[0])) == "<class 'datetime.datetime'>":
            date = content[0].strftime("%Y-%m-%d")
            dates.append(date)
            means.append(content[1])
            lows.append(content[2])
            highs.append(content[3])

    means,lows,highs = np.array(means),np.array(lows),np.array(highs)
    label = "{}, {} : {}%".format(name,dates[-1],round(means[-1]*100,3))
    print(label)

    ax.plot(dates,means,label=label)
    ax.fill_between(dates,lows,highs,alpha=0.1)

    #dates,means,lows,highs = dates[::-1],means[::-1],lows[::-1],highs[::-1]

    dates = dates[1:]
    means = new(means)
    lows = new(lows)
    highs = new(highs)

    label = "{}, {} : {}%".format(name,dates[-1],round(means[-1]*100,3))
    #print(label)

    ax_diff.plot(dates,means,label=label)
    ax_diff.fill_between(dates,lows,highs,alpha=0.2)

    ax_diff.plot(dates,np.zeros(len(dates)),'k--')
    ax.locator_params(axis='x', nbins=10)


def ONS_by_region(dataframe):
    fig,(ax1,ax2) = plt.subplots(2,1,figsize=(20,20),sharex=True)
    names = ["London","South East","East"]
    Graph_region(dataframe,['BD,BE,BF'],[ax1,ax2],names[0])
    Graph_region(dataframe,['BM,BN,BO'],[ax1,ax2],names[1])
    Graph_region(dataframe,['AU,AV,AW'],[ax1,ax2],names[2])

    ax1.set_ylim((0.0,0.05))
    ax1.set_ylabel("Fraction Infected")
    ax2.set_ylabel("Daily Change in Fraction Infected")

    ax1.legend()
    ax2.legend()

    every_nth = 4
    for n, label in enumerate(ax1.xaxis.get_ticklabels()):
        if n % every_nth != 0:
            label.set_visible(False)
    for n, label in enumerate(ax2.xaxis.get_ticklabels()):
        if n % every_nth != 0:
            label.set_visible(False)

    plt.savefig("1 ONS_regions.png")


def ONS_data():
    U1 = "https://www.ons.gov.uk/peoplepopulationandcommunity/healthandsocialcare/conditionsanddiseases/datasets/coronaviruscovid19infectionsurveydata"
    i = "https://www.ons.gov.uk/file?uri=%2fpeoplepopulationandcommunity%2fhealthandsocialcare%2fconditionsanddiseases%2fdatasets%2fcoronaviruscovid19infectionsurveydata%2f2021/"
    s = 'covid19infectionsurveydatasets'
    e = ".xlsx"

    stem = spew_text(U1,s,e)
    url = i+stem+e


    proxies = {}
    response = requests.get(url = url, proxies = proxies)
    #print("STATUS:", response.status_code)
    data = response.content
    #print(data.type)

    #f = open("out.xslx", "wb")
    #f.write(data)
    #f.close()

    #print("Warning! Scotland, wales, NI data out of whack!")

    try:
        df=pd.ExcelFile(data)
        data = df.parse(sheet_name='8a',skiprows=10,usecols = "A,B")
    except Exception as E:
        print(stem)
        print('Data Streamed', stem[-22:-10])

        print("Failed ONS parse")
        print("URL:",response.url)
        print(E)
    #print(data.type)

    ONS_by_region(df)

    stream = data.values.tolist()
    datas, dates = [],[]
    for content in stream:
        if len(dates) > 5:
            if str(type(content[0])) != "<class 'datetime.datetime'>":
                break
        if str(type(content[0])) == "<class 'datetime.datetime'>":
            date = content[0].strftime("%Y-%m-%d")
            data = content[1]
            dates.append(date)
            datas.append(data*england_pop)
    for idx in range(len(dates)):
        idx = int(idx)
    if len(dates) == len(datas):
        print(len(dates), " Extracted")
        return dates,datas
    else:
        print("Bad Data!")
        raise IndexError

def full_and_complete_analsys(dates,data,country,axislist):
    N=3 #moving average period
    data = moving_average(data,N)
    dates = dates[N-1:]

    #raw_growth_infect(dates,data,country)

    plt.figure(figsize=(15,10))
    date_of_model = '2020-12-10'

    split_point = 0
    for idx,d in enumerate(dates):
        if d == date_of_model:
            split_point = idx

    lookback = (len(dates)-split_point)
    part = len(data) - lookback

    if "USEPERCENT" is True:
        if country == 'UK':
            data = (100*data)/uk_pop
        elif country == 'England':
            data = (100*data)/england_pop
    
    dates,data = dates[part:],data[part:]

    raw_growth_infect(dates,data,country,axislist)

#Brighton and Hove
"""
bhurl = "https://cartocdn-gusc.global.ssl.fastly.net/joinzoe/api/v1/map/joinzoe@ef257c90@8e398932449a302170dd6f0080818364:1609324220703/1/attributes/23495?callback=_cdbi_layer_attributes_910841071_2"
bhurl = "https://cartocdn-gusc.global.ssl.fastly.net/joinzoe/api/v1/map/joinzoe@ef257c90@8e398932449a302170dd6f0080818364:1609583074447/1/attributes/24668?callback=_cdbi_layer_attributes_910841071_2"
proxies = {}
rs = requests.get(url = bhurl, proxies = proxies)

print("STATUS:", rs.status_code)

bh_data=rs.text

tagg = '"cases_pm_string":"'
start = bh_data.find(tagg)
end = bh_data[start:].find('",')

end = start+end
start = len(tagg) + start
#print(bh_data[start:end])

bh_data = round((float(bh_data[start:end])/1e6)*100, 3)
"""





def accumulate_numbers(week_v,dates):
    days_vacc = None
    out = []
    for day in dates:
        if day == datetime.datetime(2021, 1, 1):
            days_vacc = 0
        if days_vacc is not None:
            days_vacc += 1
            out.append(days_vacc*(week_v/7))
        else:
            out.append(0)
    return np.array(out)

def project_vacc(dates,data,daterange,n=1):

    indexlist = lambda dates, daterange : [daterange.index(date) for date in dates]

    x = indexlist(dates,daterange)

    #print(x,data)
    fit = np.polyfit(x, data, n)
    fit_fn = np.poly1d(fit)

    return fit_fn(np.arange(len(daterange))),fit[0]
        

#VACCINE
def daily_vacc(first=True):
    vurl = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nation&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22cumPeopleVaccinatedFirstDoseByPublishDate%22:%22cumPeopleVaccinatedFirstDoseByPublishDate%22,%22cumPeopleVaccinatedSecondDoseByPublishDate%22:%22cumPeopleVaccinatedSecondDoseByPublishDate%22%7D&format=json"
    uk_vaccine = get_json(vurl)['data']
    dates,datas=[],[]
    for x in range(len(uk_vaccine)):
        dates.append(uk_vaccine[x]['date'])
        if first == True:
            datas.append(uk_vaccine[x]['cumPeopleVaccinatedFirstDoseByPublishDate'])
        elif first == False:
            datas.append(uk_vaccine[x]['cumPeopleVaccinatedSecondDoseByPublishDate'])
    
    out = dict.fromkeys(dates,0)

    for day in out.keys():
        for x in range(len(uk_vaccine)):
            if uk_vaccine[x]['date'] == day:
                kah = out[day]
                if first == True:
                    kah += uk_vaccine[x]['cumPeopleVaccinatedFirstDoseByPublishDate']
                elif first == False:
                    kah += uk_vaccine[x]['cumPeopleVaccinatedSecondDoseByPublishDate']
                out[day] = kah
    
    return produce_datetimes(list(out.keys())),list(out.values())

def check_for_percent_vacc(dates_entire,data_entire, percent,pred_degree=1,raw=False):
    date_range = pd.date_range(start="2021-01-10",end="2021-12-01").to_pydatetime().tolist()
    vacc_pred,_ = project_vacc(dates_entire,data_entire,date_range,n=pred_degree)

    val = percent*uk_pop

    for idx,v in enumerate(vacc_pred):
        if v>val:
            idv = idx
            break
    
    if raw==True:
        return date_range[idv]

    return date_range[idv].strftime("%m-%d")

def check_for_percent(dates_entire,data_entire, percent):

    for idx,v in enumerate(data_entire):
        if v>percent:
            idv = idx
            break
    
    return dates_entire[idv].strftime("%m-%d")




def Vaccine(n,degree=1):
    vurl = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=overview&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22weeklyPeopleVaccinatedFirstDoseByVaccinationDate%22:%22weeklyPeopleVaccinatedFirstDoseByVaccinationDate%22,%22cumPeopleVaccinatedFirstDoseByVaccinationDate%22:%22cumPeopleVaccinatedFirstDoseByVaccinationDate%22,%22weeklyPeopleVaccinatedSecondDoseByVaccinationDate%22:%22weeklyPeopleVaccinatedSecondDoseByVaccinationDate%22,%22cumPeopleVaccinatedSecondDoseByVaccinationDate%22:%22cumPeopleVaccinatedSecondDoseByVaccinationDate%22%7D&format=json"
    uk_vaccine = get_json(vurl)['data']
    dates_v,data_vc,data_v = [],[],[]


    winning,losing = uk_pop*0.07,2e6

    for x in range(len(uk_vaccine)):
        dates_v.append(uk_vaccine[x]['date'])
        data_v.append(uk_vaccine[x]['weeklyPeopleVaccinatedFirstDoseByVaccinationDate'])
        data_vc.append(uk_vaccine[x]['cumPeopleVaccinatedFirstDoseByVaccinationDate'])

    dates_v, data_v, data_vc = dates_v[::-1],data_v[::-1],data_vc[::-1]

    fig, ax = plt.subplots(figsize=(15,15))
    ax.bar(dates_v,winning*np.ones(len(dates_v)),label = "Israel")
    ax.bar(dates_v,losing*np.ones(len(dates_v)),label = "2M per Week")
    ax.bar(dates_v,data_v,label='Actual')
    #ax.xaxis.set_major_locator(plt.MaxNLocator(7))

    llp = int((data_v[-1]/winning)*100)

    out = "Vaccinated on {} = {}\n Not Losing percentage: {}%".format(dates_v[-1],int(data_v[-1]),llp)
    #print(out)

    ax.set_title(out)
    ax.set_ylim(0,winning)
    plt.ylabel("Number Vaccinated per week")
    plt.xlabel("Week ending at")
    plt.legend()
    fig.savefig("2 finalstruggle.png")


    date_range = pd.date_range(start="2021-01-10",end="2021-07-01").to_pydatetime().tolist()

    cum_losing = accumulate_numbers(losing,date_range)
    cum_winning = accumulate_numbers(winning,date_range)
    #cum_1m = accumulate_numbers(1e6,date_range)



    fig, ax = plt.subplots(figsize=(15,15))

    dates_v = [datetime.datetime.strptime(date,'%Y-%m-%d') for date in dates_v]

    #dates_owid,data_owid = OWID_vacc()

    dates_daily,data_daily = daily_vacc(first=True)
    _,seconds_daily = daily_vacc(first=False)

    data_entire = data_daily+data_vc
    dates_entire = dates_daily+dates_v

    data_entire,dates_entire = data_daily,dates_daily

    data_entire = [x for _,x in sorted(zip(dates_entire,data_entire))]
    dates_entire = sorted(dates_entire)
   
    cutoff = dates_entire[-1] - datetime.timedelta(n)

    
    c=0
    for idx,date in enumerate(dates_entire):
        if date>cutoff:
            c = idx
            break



    dates_entire = dates_entire[c:]
    data_entire = data_entire[c:]

    diagprint(dates_entire,new(data_entire))

    o1 = "\n{} daily rate: {}".format(dates_entire[-1].strftime("%m-%d"),new(data_entire)[-1])





    ax.plot(date_range,cum_losing/uk_pop,label = "2M per Week")
    ax.plot(date_range,cum_winning/uk_pop,label = "Israel")
    #ax.plot(date_range,cum_1m/uk_pop,label = "1M per Week")
    seconds_daily = seconds_daily[::-1]

    vacc_pred,daily_rate = project_vacc(dates_entire,data_entire,date_range,n=degree)
    second_pred,_ = project_vacc(dates_entire,seconds_daily[c:],date_range,n=2)

    data_unified = data_daily[::-1]
    extra = vacc_pred[len(dates_daily):]

    for e in extra:
        data_unified.append(e)

    extra = second_pred[len(seconds_daily):]
    for e in extra:
        seconds_daily.append(e)
    
    #print(len(data_unified),len(vacc_pred),len(date_range),len(dates_daily),len(seconds_daily))

    protexs = []
    pbar=tqdm(total=len(date_range))

    phase0 = 15e6/uk_pop
    phase1 = 32e6/uk_pop

    everyone = check_for_percent_vacc(dates_entire,data_entire,1.0,pred_degree=degree)
    full_herd = check_for_percent_vacc(dates_entire,data_entire,0.8,pred_degree=degree)
    partial_herd = check_for_percent_vacc(dates_entire,data_entire,phase1,pred_degree=degree)
    prior = check_for_percent_vacc(dates_entire,data_entire,phase0,pred_degree=degree)

    over65,tdr = calculate_percent_over65(date_range,data_unified,'assumptions.xlsx')
    number_o65 = np.array(data_unified)*np.array(over65)/uk_pop
    number_u65 = np.array(data_unified)*(1-np.array(over65))/uk_pop
    number_o65_sec = np.array(seconds_daily)*np.array(over65)/uk_pop

    """
    plt.figure()
    plt.plot(date_range,over65,label='percent over65')
    plt.plot(date_range,number_o65,label='pop percent o65')
    plt.plot(date_range,number_o65_sec,label='seconds')
    print(date_range[0],date_range[-1])
    plt.legend()
    plt.savefig("age_decline_test.png")
    return
    """
    for idx,date in enumerate(date_range):
        pbar.update(1)
        if idx>30:
            o65 = over65[idx]
            onetwoweeks,twothreeweeks, rest, second = running_week_diff(date,date_range,data_unified,seconds_daily)
            protexs.append(protected_on_date(onetwoweeks,twothreeweeks, rest, second,o65))
        else:
            protexs.append(0)

    


    text = "Time for everyone:                  {}\nTime for full herd:                   {}\nTime for phase 1:                   {}\nTime for 70+ and priority:     {}".format(everyone,full_herd,partial_herd,prior)

    plt.text(0.7, 0.05, text, transform=ax.transAxes, fontsize=16, bbox=dict(facecolor='blue', alpha=0.2))

    #flu = check_for_percent(date_range,protexs,0.93)
    eighty = check_for_percent(date_range,protexs,0.8)
    seventy = check_for_percent(date_range,protexs,0.7)


    now = datetime.datetime.now()
    todays_date = datetime.datetime(now.year,now.month,now.day)

    today = date_range.index(todays_date)

    just_the_flu = (1 - float(protexs[today]))/0.07

    text = "time for 80%:    {}\n70%   {}\nprotected TODAY:{}\nO65 I/R anomaly:{}/{}\nJust the flu multiplier:{}".format(eighty,seventy,round(protexs[today],3),round(over65[today],3),round(tdr,3),round(just_the_flu,3))

    plt.text(0.1, 0.7, text, transform=ax.transAxes, fontsize=14, bbox=dict(facecolor='red', alpha=0.2))


    daily_rate = int(daily_rate)

    plt.plot(date_range,vacc_pred/uk_pop,'--', linewidth = 2.5,label = '{} day rate prediction'.format(n))

    plt.plot(date_range,protexs,label='Protected From Death, infected today')
    #plt.plot(date_range,over65,'--',label='over65%')
    #plt.plot(date_range,number_o65,'--',label = 'Percent Over 65 First dose')
    #plt.plot(date_range,number_u65,'--',label = 'Percent Under 65 First dose')
    #plt.plot(date_range,number_o65_sec,'--',label = 'over65secondpercentpop')

    #plt.plot(date_range,np.ones(len(date_range))*phase0,'--',linewidth=1.5,label="70+ and priority")
    #plt.plot(date_range,np.ones(len(date_range))*phase1,'--',linewidth=1.5,label="Phase 1")
    plt.plot(date_range,np.ones(len(date_range))*0.8,'--',linewidth=1.5,label="Full herd immunity")
    plt.plot(date_range,np.ones(len(date_range))*1.0,'--',linewidth=1.5,label="Everyone")

    ax.plot(dates_daily,np.array(data_daily)/uk_pop,'x',markersize=8, label='Daily/by report')
    ax.plot(dates_v,np.array(data_vc)/uk_pop,'x',markersize=15,label='Weekly/by actual')
    ax.plot(date_range,np.array(seconds_daily)/uk_pop,'-', label='Seconds/daily/projected')

    dates_entire,data_entire


    win_percent = vacc_pred[-1]/cum_winning[-1]

    vp = round(vacc_pred[-1]*100/uk_pop, 2)
    ax.set_ylim(0,1.0)
    
    #diagprint(dates_v,data_vc)

    out = "{} Total number/% vaccinated {} = {}%\n {} day rate final % vaccinated estimate: {}%\n{}_day daily rate estimate: {}".format(dates_entire[-1].strftime("%d-%m-%Y"),int(data_entire[-1]),round(data_entire[-1]*100/uk_pop,2),n,vp,n,daily_rate)
    out = out+o1

    print(out)
    plt.ylabel("Fraction Immunized")
    plt.title(out)
    plt.xlabel("date")
    plt.legend()
    fig.savefig("2 finalstruggle_cum{}.png".format(n))

    return win_percent


#ST HOSPITALS

def Overall_hospital():
    hurl = "https://api.coronavirus.data.gov.uk/v1/data?filters=areaType=overview&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22hospitalCases%22:%22hospitalCases%22%7D&format=json"
    uk_hospital = get_json(hurl)['data']

    maxh=0
    hn = uk_hospital[0]['hospitalCases']
    print()
    IH = "In Hospital:"+str(int(hn))+ " "+uk_hospital[0]['date']
    print(IH)
    ds=False

    dates_h,data_h = [],[]

    for x in range(len(uk_hospital)):
        nh = uk_hospital[x]['hospitalCases']
        dates_h.append(uk_hospital[x]['date'])
        data_h.append(uk_hospital[x]['hospitalCases'])
        if nh < 0.5*hn:
            if ds == False:
                doubling_time = x
                ds=True
        if nh>maxh:
            if x > len(uk_hospital)*0.8:
                maxh=nh

    FFP = "Fraction of first peak full/doublings left: "+ str(round(hn/maxh, 3))+ " " + str(round(float(np.log2(np.array(maxh/hn))),3))
    CHD = "Current Hospital Doubling: "+ str(doubling_time)
    ML = "'max(wave1) / days to equal wave 1': " + str(int(maxh)) + " " + str(int((float(np.log2(np.array(maxh/hn)))*doubling_time)))
    out = FFP+"\n"+CHD+"\n"+ML
    print(FFP)
    fig, ax = plt.subplots(figsize=(15,15))
    ax.plot(dates_h[::-1],data_h[::-1])
    ax.plot(np.arange(len(dates_h)),np.ones(len(dates_h))*maxh,'--')
    ax.xaxis.set_major_locator(plt.MaxNLocator(7))
    ax.set_title(FFP+"\n"+IH)
    plt.ylabel("In Hospital")
    fig.savefig("3 sthospitals.png")



#SYMPTOM AND ONS

def new(xvalues):
    out = np.zeros_like(xvalues)
    for idx,v in enumerate(xvalues):
        out[idx] = xvalues[idx] - xvalues[idx-1]
    return out[1:]

def raw_growth_infect(dates,data,country, axl):
    ax1,ax2 = axl
    #data = 2*np.arange(start=0,stop=len(data))
    if country == "England":
        save = "ONS_infections-England"
    else:
        save = "Symptom_infections-UK"
    
    BASE = np.arange(0, len(dates)+STEP, STEP)
    if country is not "England":
        ax1.set_xticks(BASE)
        ax2.set_xticks(BASE)
        ax2.plot(BASE,np.zeros_like(BASE),'k--',label='0')

    ax1.plot(dates,data,label=save)
    ax1.set_ylabel("Infections")

    ax2.plot(dates[1:],new(data),label=save)
    ax2.set_ylabel("Daily Change in Infections")
    ax1.legend()
    ax2.legend()
    n=1
    titl = ax1.get_title()
    titl = titl + "\n" +dates[-n] + " Total: " + str(str(round(100*data[-n]/uk_pop, 2)) +'%, ') + str(int(data[-n])) +" " +str(int(new(data)[-n]))
    ax1.set_title(titl)



def Overall_infections():
    url = 'https://joinzoe.carto.com/api/v2/sql?q=SELECT%20*%20FROM%20uk_active_cases&api_key=P6G8DgjFF94mduMkA_VroQ'
    proxies = {}
    response = requests.get(url = url, proxies = proxies)

    print("STATUS:", response.status_code)

    response=response.json()
    covid_data = response['rows']
    print('Symptom Data Online')
    dates = []
    data = np.zeros(len(covid_data))

    for idx,el in enumerate(covid_data):
        d = str(el['date'])
        d = d[:4]+'-'+d[4:6]+'-'+d[6:]
        dates.append(d)
        data[idx] = float(el['corrected_covid_positive'])

    latest = round(100*data[-1]/uk_pop, 3)

    plt.style.use('fivethirtyeight')

    print("\nONS Data...")
    dates_ONS, data_ONS = ONS_data()


    print('\nSymptom(UK)  Infections, Change:')
    for n in range(10)[1:]:
        print(dates[-n]," ", str(str(round(100*data[-n]/uk_pop, 2)) +'%'), int(data[-n]),int(new(data)[-n]))
    print('\nONS(England) Infections, Change:')
    for n in range(10)[1:]:
        print(dates_ONS[-n]," ", str(str(round(100*data_ONS[-n]/england_pop, 2)) +'%'), int(data_ONS[-n]),int(new(data_ONS)[-n]))

    fig2, (ax1, ax2) = plt.subplots(nrows=2,sharex=True,figsize=(20,10))
    #overrange = len(dates)

    full_and_complete_analsys(dates,data,'UK',[ax1, ax2])

    full_and_complete_analsys(dates_ONS,data_ONS,'England',[ax1, ax2])

    fig2.savefig('1 infections.png')

    plt.close()
    #print("\n\n LATEST:: Brighton & Hove: {} percent active cases\nwhole UK {} percent active cases".format(None,latest))

    return (dates_ONS,data_ONS),(dates,data),latest


def death_data():
    plt.style.use('fivethirtyeight')
    proj = extract_clean_doomdata('doom_data.xlsx')
    plt.figure(figsize=(15,15))
    for k in proj.keys():
        m = proj[k]
        plt.plot(m[0],m[1],'-',label=k,linewidth=1.0)

    daterange = proj['worstcase'][0]

    ensemble = np.zeros_like(np.array(daterange))

    for idx,d in enumerate(daterange):
        values_for_day = []
        for k in proj.keys():
            if k is not 'worstcase':
                m = proj[k]
                if d in m[0]:
                    daterange_idx = m[0].index(d)
                    values_for_day.append(m[1][daterange_idx])

        ensemble[idx]=np.average(values_for_day)

    plt.plot(daterange,ensemble,label='ensemble')

    durl = "https://api.coronavirus.data.gov.uk/v1/data?filters=areaType=overview&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newDeaths28DaysByPublishDate%22:%22newDeaths28DaysByPublishDate%22,%22cumDeaths28DaysByPublishDate%22:%22cumDeaths28DaysByPublishDate%22%7D&format=json"
    uk_death = get_json(durl)['data']

    dates = []
    deaths = []
    for n in uk_death:
        date = datetime.datetime.strptime(n['date'],'%Y-%m-%d')
        death = n['newDeaths28DaysByPublishDate']
        #print(date,death)
        dates.append(date)
        deaths.append(death)

    dates, deaths = dates[::-1],deaths[::-1]
    deaths = moving_average(deaths,7)
    dates = dates[6:]

    print("\nMoving AVG Deaths:")
    for n in range(5)[1:]:
        print(dates[-n].date(),round(deaths[-n],1),round(deaths[-n] - deaths[-n-1],1))

    base = daterange
    o_base = np.ones_like(base)

    #mu,std = best
    #PCR,PCRerror = mu/first, std/first
    peak1 = max(deaths[:180])

    plt.plot(base,peak1*o_base,'k--',linewidth=1.0,label='Wave 1 peak')
    #plt.plot(base,PCR*peak1*o_base,'--',linewidth=1.0,label="GJP Wave 2 mean Europe")
    #low = (PCR-PCRerror)*peak1*o_base
    #high = (PCR+PCRerror)*peak1*o_base
    #plt.fill_between(base,low.astype(int),high.astype(int),alpha=0.5)

    plt.plot(dates,deaths,label='actual')

    plt.legend()
    plt.ylabel('Deaths')
    plt.xlabel('Date')

    plt.title("Current: "+ dates[-1].strftime("%d-%m-%Y")+ " - " + str(round(deaths[-1],1)))

    plt.savefig('5 deaths.png')

    plt.xlim([datetime.date(2020,9,1), datetime.date(2021,2,1)])
    plt.savefig('5 deaths_narrow.png')


def Specific_Cases(avg_per):
    fig, ax = plt.subplots(figsize=(20,20))
    plt.title("Case growth / Positive rates")

    url1 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=utla;areaName=Kent;date%253E2020-10-02&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newCasesBySpecimenDateRollingSum%22:%22newCasesBySpecimenDateRollingSum%22,%22newCasesBySpecimenDateRollingRate%22:%22newCasesBySpecimenDateRollingRate%22,%22newCasesBySpecimenDateChange%22:%22newCasesBySpecimenDateChange%22,%22newCasesBySpecimenDateChangePercentage%22:%22newCasesBySpecimenDateChangePercentage%22%7D&format=json"
    url2 ="https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=London;date%253E2020-10-02&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newCasesBySpecimenDateRollingSum%22:%22newCasesBySpecimenDateRollingSum%22,%22newCasesBySpecimenDateRollingRate%22:%22newCasesBySpecimenDateRollingRate%22,%22newCasesBySpecimenDateChange%22:%22newCasesBySpecimenDateChange%22,%22newCasesBySpecimenDateChangePercentage%22:%22newCasesBySpecimenDateChangePercentage%22%7D&format=json"
    url3 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=East%2520of%2520England;date%253E2020-10-02&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newCasesBySpecimenDateRollingSum%22:%22newCasesBySpecimenDateRollingSum%22,%22newCasesBySpecimenDateRollingRate%22:%22newCasesBySpecimenDateRollingRate%22,%22newCasesBySpecimenDateChange%22:%22newCasesBySpecimenDateChange%22,%22newCasesBySpecimenDateChangePercentage%22:%22newCasesBySpecimenDateChangePercentage%22%7D&format=json"
    url4 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=South%2520East;date%253E2020-10-02&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newCasesBySpecimenDateRollingSum%22:%22newCasesBySpecimenDateRollingSum%22,%22newCasesBySpecimenDateRollingRate%22:%22newCasesBySpecimenDateRollingRate%22,%22newCasesBySpecimenDateChange%22:%22newCasesBySpecimenDateChange%22,%22newCasesBySpecimenDateChangePercentage%22:%22newCasesBySpecimenDateChangePercentage%22%7D&format=json"

    cases("Kent",url1,ax,'k',avg_per)
    cases("London",url2,ax,'r',avg_per)
    cases("East of England",url3,ax,'g',avg_per)
    cases("South East",url4,ax,'b',avg_per)

    url1 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=utla;areaName=Kent&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22uniqueCasePositivityBySpecimenDateRollingSum%22:%22uniqueCasePositivityBySpecimenDateRollingSum%22,%22uniquePeopleTestedBySpecimenDateRollingSum%22:%22uniquePeopleTestedBySpecimenDateRollingSum%22%7D&format=json"
    url2 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=London&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22uniqueCasePositivityBySpecimenDateRollingSum%22:%22uniqueCasePositivityBySpecimenDateRollingSum%22,%22uniquePeopleTestedBySpecimenDateRollingSum%22:%22uniquePeopleTestedBySpecimenDateRollingSum%22%7D&format=json"
    url3 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=East%2520of%2520England&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22uniqueCasePositivityBySpecimenDateRollingSum%22:%22uniqueCasePositivityBySpecimenDateRollingSum%22,%22uniquePeopleTestedBySpecimenDateRollingSum%22:%22uniquePeopleTestedBySpecimenDateRollingSum%22%7D&format=json"
    url4= "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=region;areaName=South%2520East&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22uniqueCasePositivityBySpecimenDateRollingSum%22:%22uniqueCasePositivityBySpecimenDateRollingSum%22,%22uniquePeopleTestedBySpecimenDateRollingSum%22:%22uniquePeopleTestedBySpecimenDateRollingSum%22%7D&format=json"

    positive("Kent",url1,ax,'k',avg_per)
    positive("London",url2,ax,'r',avg_per)
    positive("East of England",url3,ax,'g',avg_per)
    positive("South East",url4,ax,'b',avg_per)


    ax.xaxis.set_major_locator(plt.MaxNLocator(12))

    plt.xlim(datetime.datetime(2020,12,1),datetime.datetime(2021,4,15))
    date_range = pd.date_range(start="2020-12-1",end="2021-04-01").to_pydatetime().tolist()


    plt.plot(date_range,np.zeros(len(date_range)),'k--',label="0")
    plt.legend()


    plt.savefig("4 cases {}.png".format(avg_per))


def Specific_Hospitals():
    fig, ax = plt.subplots(figsize=(20,20))
    plt.title("In-Hospital Change")
    date_range = pd.date_range(start="2020-12-1",end="2021-04-01").to_pydatetime().tolist()

    url2 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=London&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newAdmissions%22:%22newAdmissions%22,%22cumAdmissions%22:%22cumAdmissions%22%7D&format=json"
    url3 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=East%2520of%2520England&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newAdmissions%22:%22newAdmissions%22,%22cumAdmissions%22:%22cumAdmissions%22%7D&format=json"
    url4 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=South%2520East&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22newAdmissions%22:%22newAdmissions%22,%22cumAdmissions%22:%22cumAdmissions%22%7D&format=json"

    hospitalsadmissions("London",url2,ax,'r--')
    hospitalsadmissions("East of England",url3,ax,'g--')
    hospitalsadmissions("South East",url4,ax,'b--')

    url2 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=London&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22hospitalCases%22:%22hospitalCases%22%7D&format=json"
    url3 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=East%2520of%2520England&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22hospitalCases%22:%22hospitalCases%22%7D&format=json"
    url4 = "https://coronavirus.data.gov.uk/api/v1/data?filters=areaType=nhsregion;areaName=South%2520East&structure=%7B%22areaType%22:%22areaType%22,%22areaName%22:%22areaName%22,%22areaCode%22:%22areaCode%22,%22date%22:%22date%22,%22hospitalCases%22:%22hospitalCases%22%7D&format=json"

    hospitalschange("London",url2,ax,'r')
    hospitalschange("East of England",url3,ax,'g')
    hospitalschange("South East",url4,ax,'b')

    plt.plot(date_range,np.ones(len(date_range)),'k--',label="0")
    plt.xlim(datetime.datetime(2020,12,8),datetime.datetime(2021,4,1))
    plt.legend()

    plt.title("Change In hospital")
    #plt.show()
    plt.savefig("3 regions_hospitals.png")

#print("Local/National ratio :: {}".format(round(bh_data/latest, 2)))

def Symptom_pdf():
    url = 'https://covid-assets.joinzoe.com/latest/covid_symptom_study_report.pdf'
    response = requests.get(url = url, proxies = {})
    data = response.content

    symptom_pdf = open('ss.pdf', 'wb')
    symptom_pdf.write(data)
    symptom_pdf.close()
    response.headers

    viewer = SimplePDFViewer(data)
    viewer.navigate(7)
    viewer.render()
    for k,v in viewer.canvas.images.items():
        image = v.to_Pillow()
        name = 'ss'+k+'.png'
        image.save(name)
        print(name)



def running_week_diff(now,dates,first_doses,second_doses):
    dates,first_doses,second_doses = dates[::-1],first_doses[::-1],second_doses[::-1]
    for idx,d in enumerate(dates):
        if d == now:
            shift = idx
    one,two,three = 7+shift,14+shift,21+shift


    onetwoweeks = first_doses[one]-first_doses[two]
    twothreeweeks = first_doses[two]-first_doses[three]
    rest = first_doses[three]

    rest = first_doses[three]


    second = second_doses[one-1]    
    percentpop = lambda number : round(100*number/uk_pop,3)
   

    return percentpop(onetwoweeks),percentpop(twothreeweeks), percentpop(rest), percentpop(second)

def obtain_rob():
    """
    rourl = "https://docs.google.com/spreadsheets/d/1y6rExrXJKUz2ljuzAY3kALkpbAAe-mGzxCe80K45OIQ/edit#gid=344543969"
    proxies = {}
    response = requests.get(url = rourl, proxies = proxies)

    print("STATUS:", response.status_code)

    data=response.content

    f = open("conc.xlsx", "wb")
    f.write(data)
    f.close()


    wb = openpyxl.load_workbook(filename='conc.xlsx')
    print("Infections,Hospitals,Deaths:")
    print(wb._sheets[0]['E4'])
    print(wb._sheets[0]['E5'])
    print(wb._sheets[0]['E6'])
    """



    rurl = 'https://docs.google.com/spreadsheets/d/1y6rExrXJKUz2ljuzAY3kALkpbAAe-mGzxCe80K45OIQ/export?gid=0&format=xlsx'

    proxies = {}
    response = requests.get(url = rurl, proxies = proxies)

    print("STATUS:", response.status_code)

    data=response.content

    f = open("assumptions.xlsx", "wb")
    f.write(data)
    f.close()


    wb = openpyxl.load_workbook(filename='assumptions.xlsx')
    
    wb._sheets[0]['Y3'] = 1
    wb._sheets[0]['D2'] = 1
    wb._sheets[0]['D3'] = 1
    wb._sheets[0]['D4'] = 1
    wb._sheets[0]['D5'] = 1

    wb._sheets[0]['D9'] = 1
    wb._sheets[0]['D10'] = 1
    wb._sheets[0]['D11'] = 1
    wb._sheets[0]['D12'] = 1

    wb._sheets[0]['D16'] = 1
    wb._sheets[0]['D17'] = 1
    wb._sheets[0]['D18'] = 1
    wb._sheets[0]['D19'] = 1

    wb.save(filename = 'assump.xlsx')
    wb.close()


    print("Assump refreshed")
    return data





def calculate_percent_over65(daterange,data_unified,filename):
    data_unified = np.array(data_unified)/uk_pop
    everyone_date = datetime.datetime(2021,6,1)
    everyone_ratio = read_single_value(filename,['V',3])
    now = datetime.datetime.now()

    todays_date = datetime.datetime(now.year,now.month,now.day)
    todays_ratio = read_single_value(filename,['Y',3])

    dates = [todays_date,everyone_date]
    data = [todays_ratio,everyone_ratio]

    indexlist = lambda dates, daterange : [daterange.index(date) for date in dates]

    x = indexlist(dates,daterange)

    #print(x,data)
    fit = np.polyfit(x, data, 1)
    fit_fn = np.poly1d(fit)

    percents = fit_fn(np.arange(len(daterange)))

    percent_overall_over65 = np.array(data_unified)*np.array(percents)

    exceed_error = percent_overall_over65 > everyone_ratio
    make_true=False
    for idx,E in enumerate(exceed_error):
        if E == True:
            make_true=True
        if make_true == True:
            exceed_error[idx] = True

    corrected = []
    for idx in range(len(percents)):
        wholepop_under65_now = data_unified[idx] - everyone_ratio
        percent_now = 1 - (wholepop_under65_now/data_unified[idx])
        corrected.append(percent_now)

    raw_old = np.invert(exceed_error)

    over65 = (exceed_error*np.array(corrected)) + (raw_old*percents)
    
    return over65,todays_ratio


def read_single_value(filename,CR,sheet='Data and assumptions'):
    out = pd.read_excel(filename, sheet_name=sheet,index_col=None, usecols = CR[0],skiprows=CR[1]-1,nrows=1)
    returning = float(out.columns.values[0])
    return returning


def protected_on_date(onetwoweeks,twothreeweeks, rest, second, o65,filename = 'assump.xlsx'):
    over = read_single_value(filename,['S',3])
    under = 1-over

    
    under_over = [under,over]
    wb = openpyxl.load_workbook(filename=filename)

    first_doses_oneplus = (onetwoweeks+twothreeweeks+rest)/100
    second_doses_oneplus = second/100

    corrector = (first_doses_oneplus-second_doses_oneplus)/first_doses_oneplus
    
    wb._sheets[0]['Y3'] = o65
    wb._sheets[0]['D2'] = corrector*onetwoweeks/100
    wb._sheets[0]['D3'] = corrector*twothreeweeks/100
    wb._sheets[0]['D4'] = corrector*rest/100
    wb._sheets[0]['D5'] = second/100
    wb.save(filename = 'A0.xlsx')
    wb.close()


    effect_model = formulas.ExcelModel().loads('A0.xlsx').finish(circular=True)
    result = effect_model.calculate(outputs=["'[A0.xlsx]DATA AND ASSUMPTIONS'!AK2","'[A0.xlsx]DATA AND ASSUMPTIONS'!AK3"])

    y=float(result["'[A0.xlsx]DATA AND ASSUMPTIONS'!AK2"].value[0])
    o=float(result["'[A0.xlsx]DATA AND ASSUMPTIONS'!AK3"].value[0])

    if y>1.00:
        y=1.00
        print("Anomaly!")
        print("o65:{}, weeks:{},{},{},{}\n:y{} o{}".format(o65,onetwoweeks/100,twothreeweeks/100,rest/100,second/100,y,o))
    
    if o>1.00:
        y=1.00
        print("Anomaly!")
        print("o65:{}, weeks:{},{},{},{}\n:y{} o{}".format(o65,onetwoweeks/100,twothreeweeks/100,rest/100,second/100,y,o))

    
    protected = (y*under_over[0]) + (o*under_over[1])


    return protected

if __name__ == "__main__":
    """
    print("\n\n 0 Symptom Survey Breakdown by region")
    Symptom_pdf()
    
    print("\n\n 1 Extrinsic Information on Infections")
    try:
        ONS, symptom,latest = Overall_infections()
        ONS = (produce_datetimes(ONS[0]),ONS[1])
        symptom = (produce_datetimes(symptom[0]),symptom[1])
    except Exception as E:
        print(E)
        print("Failed stream ONS/Symptom")

    print("\n\n 2 Vaccinations - the Final Struggle")

    """
    #obtain_rob()
    Vaccine(20)
    """
    print("\n\n 3 Hospitals")
    Overall_hospital()
    Specific_Hospitals()

    print("\n\n 4 Growth Rates")
    Specific_Cases(1)
    #print("\n7 Day Average:")
    #Specific_Cases(7)

    print("\n\n 5 Deaths")
    death_data()

    print("\n\n 6 Model Projections")
    SE_hospitals(3,vaccine_ratio)

    print("Latest, UK {}% symptomatic".format(latest))
    """